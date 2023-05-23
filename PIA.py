import sqlite3 
from sqlite3 import Error
import os
import openpyxl
import csv
import sys
import datetime

libros = dict()

try:
    with open("libros.csv", 'r', newline="") as archivo:
        lector = csv.reader(archivo)
        next(lector)
        for identificador, titulo, autor, genero, añoPublic, isbn, fechaAdqui in lector:
            libros[int(identificador)] = (titulo, autor, genero, int(añoPublic), isbn, fechaAdqui)
except:
    print()
    print('No se ha encontrado ningún archivo previo de guardado')
    print('**SE HA GENERADO UN ARCHIVO CSV EN BLANCO**')
    archivo = open("libros.csv", "w", newline="")
    grabador = csv.writer(archivo)
    grabador.writerow(("Identificador", "Título", "Autor", 'Género', 'Año de publicación', "ISBN", 'Fecha de adquisición'))
    archivo.close()

try:
    with sqlite3.connect("Biblioteca.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS libro (identificador INTEGER PRIMARY KEY, titulo TEXT, autor_id INTEGER, \
         genero_id INTEGER, año_publicacion INTEGER, isbn TEXT, fecha_adquisicion timestamp, FOREIGN KEY(autor_id) \
         REFERENCES autor(clave), FOREIGN KEY(genero_id) REFERENCES genero(clave));")
        print("Tabla 'libro' creada exitosamente")
except Error as e:
    print(e)
except Exception as e:
    print(e)

def tabla_autor(nombre, apellido):
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("CREATE TABLE IF NOT EXISTS autor (clave INTEGER PRIMARY KEY, nombre TEXT NOT NULL, \
             apellido TEXT NOT NULL);")
            print("Tabla 'autor' creada exitosamente")
    except Error as ex:
        print(ex)
    except Exception as ex:
        print(ex)

    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            valores = (nombre, apellido)
            mi_cursor.execute("INSERT INTO autor (nombre, apellido) VALUES(?,?)", valores)
            autor_id = mi_cursor.lastrowid
            print(f"El autor ha sido registrado con el ID {autor_id}")
            return autor_id
    except Error as ex:
        print(ex)
    except Exception as ex:
        print(ex)

def tabla_genero(genero):
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("CREATE TABLE IF NOT EXISTS genero (clave INTEGER PRIMARY KEY, genero TEXT NOT NULL);")
            print("Tabla 'genero' creada exitosamente")
    except Error as ex:
        print(ex)
    except Exception as ex:
        print(ex)

    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            valores = (genero,)
            mi_cursor.execute("INSERT INTO genero (genero) VALUES(?)", valores)
            genero_id = mi_cursor.lastrowid
            print(f"El género ha sido registrado con el ID {genero_id}")
            return genero_id
    except Error as ex:
        print(ex)
    except Exception as ex:
        print(ex)

def RegistrarNuevoEjemplar():
    print("Registrar nuevo ejemplar")
    print("------------------------")
    titulo = input("Título del libro: ")
    autor_nombre = input("Nombre del autor: ")
    autor_apellido = input("Apellido del autor: ")
    autor_id = tabla_autor(autor_nombre, autor_apellido)
    genero = input("Género: ")
    genero_id = tabla_genero(genero)
    año_publicacion = int(input("Año de publicación: "))
    isbn = input("ISBN: ")
    fecha_adquisicion = input("Fecha de adquisición (dd/mm/aaaa): ")


    try:
        fecha_adquisicion = datetime.datetime.strptime(fecha_adquisicion, "%d/%m/%Y")
    except ValueError:
        print("Formato de fecha incorrecto. Asegúrate de ingresar la fecha en el formato dd/mm/aaaa.")
        return


    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            valores = (titulo, autor_id, genero_id, año_publicacion, isbn, fecha_adquisicion)
            mi_cursor.execute("INSERT INTO libro (titulo, autor_id, genero_id, año_publicacion, isbn, fecha_adquisicion) \
                VALUES(?, ?, ?, ?, ?, ?)", valores)
            print("El nuevo ejemplar ha sido registrado correctamente.")
    except Error as ex:
        print(ex)
    except Exception as ex:
        print(ex)

    opcion = input("¿Deseas registrar otro ejemplar? (S/N): ")
    if opcion.lower() == "s":
        RegistrarNuevoEjemplar()

def MostrarCatalogoCompleto():
    try:
        with sqlite3.connect("Biblioteca.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT l.identificador, l.titulo, a.nombre, g.genero, l.año_publicacion, l.isbn, l.fecha_adquisicion \
                FROM libro AS l \
                JOIN autor AS a ON l.autor_id = a.clave \
                JOIN genero AS g ON l.genero_id = g.clave")
            registros = mi_cursor.fetchall()

            if registros:
                print(f"{'ID'}\t|{'Titulo':15}|{'Autor':10}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'Fecha Adquisición'}")
                print("*" * 90)
                for registro in registros:
                    identificador, titulo, autor, genero, año_publicacion, isbn, fecha_adquisicion = registro
                    print(f"{identificador}\t|{titulo:<15}|{autor:<10}|{genero:<10}|{año_publicacion:<18}|{isbn:<13}|{fecha_adquisicion.date()}")
            else:
                print("No se encontraron registros en la base de datos.")
    except sqlite3.Error as e:
        print(e)
    except Exception as e:
        print(e)

    opcion = input("¿Deseas exportar el catálogo a un archivo? (S/N): ")
    if opcion.lower() == "s":
        tipo_exportacion = input("Selecciona el formato de exportación (CSV/Excel): ")
        if tipo_exportacion.lower() == "csv":
            exportar_catalogo_csv()
        elif tipo_exportacion.lower() == "excel":
            exportar_catalogo_excel()
        else:
            print("Opción de exportación no válida.")

def exportar_catalogo_csv():
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT l.identificador, l.titulo, a.nombre, g.genero, l.año_publicacion, l.isbn, l.fecha_adquisicion \
                FROM libro AS l \
                JOIN autor AS a ON l.autor_id = a.clave \
                JOIN genero AS g ON l.genero_id = g.clave")
            registros = mi_cursor.fetchall()

            if registros:
                with open("catalogo.csv", "w", newline="") as archivo_csv:
                    escritor = csv.writer(archivo_csv)
                    escritor.writerow(["Identificador", "Título", "Autor", "Género", "Año de publicación", "ISBN", "Fecha de adquisición"])
                    escritor.writerows(registros)
                print("El catálogo se ha exportado correctamente a 'catalogo.csv'")
            else:
                print("No se encontraron registros en la base de datos.")
    except sqlite3.Error as e:
        print(e)
    except Exception as e:
        print(e)

def exportar_catalogo_excel():
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT l.identificador, l.titulo, a.nombre, g.genero, l.año_publicacion, l.isbn, l.fecha_adquisicion \
                FROM libro AS l \
                JOIN autor AS a ON l.autor_id = a.clave \
                JOIN genero AS g ON l.genero_id = g.clave")
            registros = mi_cursor.fetchall()

            if registros:
                libro_excel = openpyxl.Workbook()
                hoja = libro_excel.active
                hoja.append(["Identificador", "Título", "Autor", "Género", "Año de publicación", "ISBN", "Fecha de adquisición"])
                for registro in registros:
                    hoja.append(registro)
                libro_excel.save("catalogo.xlsx")
                print("El catálogo se ha exportado correctamente a 'catalogo.xlsx'")
            else:
                print("No se encontraron registros en la base de datos.")
    except sqlite3.Error as e:
        print(e)
    except Exception as e:
        print(e)

def ReportePorAutor():
    try:
        with sqlite3.connect("Biblioteca.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM autor ORDER BY nombre")
            autores = mi_cursor.fetchall()

            if autores:
                print("Autores registrados:")
                print("Clave\tNombre\tApellido")
                print("*" * 30)
                for clave, nombre, apellido in autores:
                    print(f"{clave}\t{nombre}\t{apellido}")
                print()

                autor_nombre = input("Ingrese el nombre del autor: ")

                mi_cursor.execute("SELECT l.identificador, l.titulo, a.nombre, g.genero, l.año_publicacion, l.isbn, l.fecha_adquisicion \
                    FROM libro AS l \
                    JOIN autor AS a ON l.autor_id = a.clave \
                    JOIN genero AS g ON l.genero_id = g.clave \
                    WHERE a.nombre = ?", (autor_nombre,))
                registros = mi_cursor.fetchall()

                if registros:
                    print(f"{'ID'}\t|{'Titulo':15}|{'Autor':10}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'Fecha Adquisición'}")
                    print("*" * 90)
                    for registro in registros:
                        identificador, titulo, autor, genero, año_publicacion, isbn, fecha_adquisicion = registro
                        print(f"{identificador}\t|{titulo:<15}|{autor:<10}|{genero:<10}|{año_publicacion:<18}|{isbn:<13}|{fecha_adquisicion.date()}")

                    opcion_exportar = input("¿Desea exportar los registros? (S/N): ")
                    if opcion_exportar.lower() == "s":
                        tipo_exportacion = input("Seleccione el formato de exportación (CSV/Excel): ")
                        if tipo_exportacion.lower() == "csv":
                            exportar_autor_csv(registros)
                        elif tipo_exportacion.lower() == "excel":
                            exportar_autor_excel(registros)
                        else:
                            print("Opción de exportación no válida.")
                else:
                    print("No se encontraron registros para el autor especificado.")
            else:
                print("No hay autores registrados en la base de datos.")
    except sqlite3.Error as e:
        print(e)
    except Exception as e:
        print(e)

def exportar_autor_csv(registros):
    try:
        with open("registros_autor.csv", "w", newline="") as archivo_csv:
            escritor = csv.writer(archivo_csv)
            escritor.writerow(["Identificador", "Título", "Autor", "Género", "Año de publicación", "ISBN", "Fecha de adquisición"])
            escritor.writerows(registros)
        print("Los registros se han exportado correctamente a 'registros_autor.csv'")
    except Exception as e:
        print("Error al exportar los registros a CSV:", str(e))

def exportar_autor_excel(registros):
    try:
        libro_excel = openpyxl.Workbook()
        hoja = libro_excel.active
        hoja.append(["Identificador", "Título", "Autor", "Género", "Año de publicación", "ISBN", "Fecha de adquisición"])
        for registro in registros:
            hoja.append(registro)
        libro_excel.save("registros_autor.xlsx")
        print("Los registros se han exportado correctamente a 'registros_autor.xlsx'")
    except Exception as e:
        print("Error al exportar los registros a Excel:", str(e))

def Reporte_generos():
    try:
        with sqlite3.connect("Biblioteca.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM genero ORDER BY genero")
            generos = mi_cursor.fetchall()

            if generos:
                print("Géneros registrados:")
                print("Clave\tGénero")
                print("*" * 30)
                for clave, genero in generos:
                    print(f"{clave}\t{genero}")
                print()

                genero_elegido = input("Ingrese el nombre del género: ")

                mi_cursor.execute("SELECT l.identificador, l.titulo, a.nombre, g.genero, l.año_publicacion, l.isbn, l.fecha_adquisicion \
                    FROM libro AS l \
                    JOIN autor AS a ON l.autor_id = a.clave \
                    JOIN genero AS g ON l.genero_id = g.clave \
                    WHERE g.genero = ?", (genero_elegido,))
                registros = mi_cursor.fetchall()

                if registros:
                    print(f"{'ID'}\t|{'Titulo':15}|{'Autor':10}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'Fecha Adquisición'}")
                    print("*" * 90)
                    for registro in registros:
                        identificador, titulo, autor, genero, año_publicacion, isbn, fecha_adquisicion = registro
                        print(f"{identificador}\t|{titulo:<15}|{autor:<10}|{genero:<10}|{año_publicacion:<18}|{isbn:<13}|{fecha_adquisicion.date()}")

                    opcion_exportar = input("¿Desea exportar los registros? (S/N): ")
                    if opcion_exportar.lower() == "s":
                        tipo_exportacion = input("Seleccione el formato de exportación (CSV/Excel): ")
                        if tipo_exportacion.lower() == "csv":
                            exportar_genero_csv(registros)
                        elif tipo_exportacion.lower() == "excel":
                            exportar_genero_excel(registros)
                        else:
                            print("Opción de exportación no válida.")
                else:
                    print("No se encontraron registros para el género especificado.")
            else:
                print("No hay géneros registrados en la base de datos.")
    except sqlite3.Error as e:
        print(e)
    except Exception as e:
        print(e)

def exportar_genero_csv(registros):
    try:
        with open("registros_genero.csv", "w", newline="") as archivo_csv:
            escritor = csv.writer(archivo_csv)
            escritor.writerow(["Identificador", "Título", "Autor", "Género", "Año de publicación", "ISBN", "Fecha de adquisición"])
            escritor.writerows(registros)
        print("Los registros se han exportado correctamente a 'registros_genero.csv'")
    except Exception as e:
        print("Error al exportar los registros a CSV:", str(e))

def exportar_genero_excel(registros):
    try:
        libro_excel = openpyxl.Workbook()
        hoja = libro_excel.active
        hoja.append(["Identificador", "Título", "Autor", "Género", "Año de publicación", "ISBN", "Fecha de adquisición"])
        for registro in registros:
            hoja.append(registro)
        libro_excel.save("registros_genero.xlsx")
        print("Los registros se han exportado correctamente a 'registros_genero.xlsx'")
    except Exception as e:
        print("Error al exportar los registros a Excel:", str(e))

def Reporte_años_publicacion():
    try:
        with sqlite3.connect("Biblioteca.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT DISTINCT año_publicacion FROM libro ORDER BY año_publicacion")
            años_publicacion = mi_cursor.fetchall()

            if años_publicacion:
                print("Años de publicación registrados:")
                print("Año")
                print("*" * 15)
                for año in años_publicacion:
                    print(año[0])
                print()

                año_elegido = input("Ingrese el año de publicación: ")

                mi_cursor.execute("SELECT l.identificador, l.titulo, a.nombre, g.genero, l.año_publicacion, l.isbn, l.fecha_adquisicion \
                    FROM libro AS l \
                    JOIN autor AS a ON l.autor_id = a.clave \
                    JOIN genero AS g ON l.genero_id = g.clave \
                    WHERE l.año_publicacion = ?", (año_elegido,))
                registros = mi_cursor.fetchall()

                if registros:
                    print(f"{'ID'}\t|{'Titulo':15}|{'Autor':10}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'Fecha Adquisición'}")
                    print("*" * 90)
                    for registro in registros:
                      identificador, titulo, autor, genero, año_publicacion, isbn, fecha_adquisicion = registro
                      print(f"{identificador}\t|{titulo:<15}|{autor:<10}|{genero:<10}|{año_publicacion:<18}|{isbn:<13}|{fecha_adquisicion.date()}")

                    opcion_exportar = input("¿Desea exportar los registros? (S/N): ")
                    if opcion_exportar.lower() == "s":
                        tipo_exportacion = input("Seleccione el formato de exportación (CSV/Excel): ")
                        if tipo_exportacion.lower() == "csv":
                            exportar_año_csv(registros)
                        elif tipo_exportacion.lower() == "excel":
                            exportar_año_excel(registros)
                        else:
                            print("Opción de exportación no válida.")
                else:
                    print("No se encontraron registros para el año de publicación especificado.")
            else:
                print("No hay años de publicación registrados en la base de datos.")
    except sqlite3.Error as e:
        print(e)
    except Exception as e:
        print(e)

def exportar_año_csv(registros):
    try:
        with open("registros_anos_publicacion.csv", "w", newline="") as archivo_csv:
            escritor = csv.writer(archivo_csv)
            escritor.writerow(["Identificador", "Título", "Autor", "Género", "Año de publicación", "ISBN", "Fecha de adquisición"])
            escritor.writerows(registros)
        print("Los registros se han exportado correctamente a 'registros_anos_publicacion.csv'")
    except Exception as e:
        print("Error al exportar los registros a CSV:", str(e))

def exportar_año_excel(registros):
    try:
        libro_excel = openpyxl.Workbook()
        hoja = libro_excel.active
        hoja.append(["Identificador", "Título", "Autor", "Género", "Año de publicación", "ISBN", "Fecha de adquisición"])
        for registro in registros:
            hoja.append(registro)
        libro_excel.save("registros_anos_publicacion.xlsx")
        print("Los registros se han exportado correctamente a 'registros_anos_publicacion.xlsx'")
    except Exception as e:
        print("Error al exportar los registros a Excel:", str(e))

def BusquedaPorTitulo():
    try:
        with sqlite3.connect("Biblioteca.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT titulo FROM libro ")
            titulos = mi_cursor.fetchall()
            if titulos:
                print("Títulos registrados:")
                print("*" * 30)
                for titulo in titulos:
                    print(titulo[0])

            titulo = input("Ingrese el título del libro: ")
            mi_cursor.execute("SELECT l.identificador, l.titulo, a.nombre, g.genero, l.año_publicacion, l.isbn, l.fecha_adquisicion \
                FROM libro AS l \
                JOIN autor AS a ON l.autor_id = a.clave \
                JOIN genero AS g ON l.genero_id = g.clave \
                WHERE l.titulo LIKE ?", ('%' + titulo + '%',))
            registros = mi_cursor.fetchall()

            if registros:
                print("Registros encontrados:")
                print(f"{'ID'}\t|{'Titulo':15}|{'Autor':10}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'Fecha Adquisición'}")
                print("*" * 90)
                for registro in registros:
                  identificador, titulo, autor, genero, año_publicacion, isbn, fecha_adquisicion = registro
                  print(f"{identificador}\t|{titulo:<15}|{autor:<10}|{genero:<10}|{año_publicacion:<18}|{isbn:<13}|{fecha_adquisicion.date()}")
    except Exception:
          print("\nDebes ingresar un valor entero. Por favor inténtalo de nuevo.")

def BusquedaPorISBN():
    try:
        with sqlite3.connect("Biblioteca.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT isbn FROM libro ")
            isbns = mi_cursor.fetchall()
            if isbns:
                print("isbn registrados:")
                print("*" * 30)
                for isbn in isbns:
                    print(isbn[0])

            isbn_elegido = input("Ingrese el isbndel libro: ")
            mi_cursor.execute("SELECT l.identificador, l.titulo, a.nombre, g.genero, l.año_publicacion, l.isbn, l.fecha_adquisicion \
              FROM libro AS l \
              JOIN autor AS a ON l.autor_id = a.clave \
              JOIN genero AS g ON l.genero_id = g.clave \
              WHERE l.isbn = ?", (isbn_elegido,))
            registros = mi_cursor.fetchall()

            if registros:
                print("Registros encontrados:")
                print(f"{'ID'}\t|{'Titulo':15}|{'Autor':10}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'Fecha Adquisición'}")
                print("*" * 90)
                for registro in registros:
                  identificador, titulo, autor, genero, año_publicacion, isbn, fecha_adquisicion = registro
                  print(f"{identificador}\t|{titulo:<15}|{autor:<10}|{genero:<10}|{año_publicacion:<18}|{isbn:<13}|{fecha_adquisicion.date()}")
    except Exception:
          print("\nDebes ingresar un valor entero. Por favor inténtalo de nuevo.")


def ConsultaYReportes():
    while True:
        print()
        print("*********CONSULTA Y REPORTES****************")
        
        print("*1* Consulta de titulo y ISBN")
        print("*2* Reportes")
        print("*3* Volver al menu principal")
        try:
          eleccion=int(input("Elige una opcion: "))
          if eleccion==1:
            TituloYIsbn()
          if eleccion==2:
            Reportes()
          if eleccion==3:
              break
          else:
            print("\nOpción inválida. Por favor eliga una opción válida.")
        except Exception:
          print("\nDebes ingresar un valor entero. Por favor inténtalo de nuevo.")

def TituloYIsbn():
    while True:
        print()
        print("**********Consulta por titulo y ISBN****")
        print()
        print("*1* busqueda por Titulo")
        print("*2* Busqueda por ISBN")
        print("*3* Volver al menu principal")
        try:
          eleccion=int(input("Elige una opcion: "))
          if eleccion==1:
            BusquedaPorTitulo()
          if eleccion==2:
            BusquedaPorISBN()
          if eleccion==3:
            break
          else:
            print("\nOpción inválida. Por favor eliga una opción válida.")
        except Exception:
          print("\nDebes ingresar un valor entero. Por favor inténtalo de nuevo.")

def Reportes():
    while True:
        print()
        print("*****Reportes*****")
        print("*1* Catalago completo")
        print("*2* Reporte por autor")
        print('*3* Reporte por genero')
        print('*4* Reporte por año de publicacion')
        print('*5* Regresar al menu anterior')
        try:
          eleccion=int(input("Selecciona una opcion: "))
          if eleccion==1:
            MostrarCatalogoCompleto()
          if eleccion==2:
            ReportePorAutor()
          if eleccion==3:
            Reporte_generos()
          if eleccion==4:
            Reporte_años_publicacion()
          if eleccion==5:
              break
          else:
            print("\nOpción inválida. Por favor eliga una opción válida.")
        except Exception:
          print("\nDebes ingresar un valor entero. Por favor inténtalo de nuevo.")
def Menu():
    while True:
        print()
        print("***********BIBLIOTECA*************")
        print()
        print("*1* Registrar nuevo ejemplar")
        print("*2* Consultas y reportes")
        print("*3* Registrar autor")
        print("*4* Registrar genero")
        print("*5* Salir")
        print()
        var_elect=input("Ingrese un numero: ")
        if var_elect=="1":
            RegistrarNuevoEjemplar()
        if var_elect=="2":
            ConsultaYReportes()
        elif var_elect=="3":
            try:
                with sqlite3.connect("Autores.db") as conn:
                  mi_cursor = conn.cursor()
                  mi_cursor.execute("SELECT * FROM autor ORDER BY nombre")
                  registros = mi_cursor.fetchall()
                  if registros:
                    print("Claves\tnombre\tApellido")
                    print("*" * 30)
                    for clave,nombre,apellido in registros:
                      print(f"{clave:^6}\t{nombre:<10}\t{apellido:<10}")
                  else:
                        print("No se encontraron registros en la respuesta")
            except Error as ex:
                print (ex)
            except Exception:
                print (ex)
        elif var_elect=="4":
            tabla_genero(genero)
            try:
                with sqlite3.connect("Generos.db") as conn:
                  mi_cursor = conn.cursor()
                  mi_cursor.execute("SELECT * FROM genero ORDER BY genero")
                  registros = mi_cursor.fetchall()
                  if registros:
                    print("Claves\tGenero")
                    print("*" * 30)
                    for clave,genero, in registros:
                      print(f"{clave:^6}\t{genero:<10}")
                  else:
                        print("No se encontraron registros en la respuesta")
            except Error as ex:
                print (ex)
            except Exception:
                print (ex)
            try:
                with sqlite3.connect("Generos.db") as conn:
                  mi_cursor = conn.cursor()
                  mi_cursor.execute("SELECT * FROM genero ORDER BY genero")
                  registros = mi_cursor.fetchall()
                  if registros:
                    print("Claves\tGenero")
                    print("*" * 30)
                    for clave,genero in registros:
                      print(f"{clave:^6}\t{genero:<10}")
                  else:
                        print("No se encontraron registros en la respuesta")
            except Error as ex:
                print (ex)
            except Exception:
                print (ex)
        elif var_elect=="5":
            try:
              with sqlite3.connect("Biblioteca.db") as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT l.identificador, l.titulo, a.nombre, g.genero, l.año_publicacion, l.isbn, l.fecha_adquisicion \
                  FROM libro AS l \
                  JOIN autor AS a ON l.autor_id = a.clave \
                  JOIN genero AS g ON l.genero_id = g.clave")
              registros = mi_cursor.fetchall()

              if registros:
                  with open("libros.csv", "w", newline="") as archivo_csv:
                      escritor = csv.writer(archivo_csv)
                      escritor.writerow(["Identificador", "Título", "Autor", "Género", "Año de publicación", "ISBN", "Fecha de adquisición"])
                      escritor.writerows(registros)
                  print("Ha salido del programa")
                  break
              else:
                  print("No se encontraron registros en la base de datos.")
            except sqlite3.Error as e:
              print(e)
            except Exception as e:
              print(e)
            else:
              print("\nOpción inválida. Por favor eliga una opción válida.")
            
Menu()